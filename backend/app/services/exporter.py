import io
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ExportService:
    @staticmethod
    def generate_pdf_report(data: dict) -> bytes:
        """
        Generates a professional PDF report from the analysis JSON.
        """
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=letter, 
            rightMargin=54, 
            leftMargin=54, 
            topMargin=54, 
            bottomMargin=54
        )
        
        styles = getSampleStyleSheet()
        
        # Define Custom Styles for Premium Appearance
        title_style = ParagraphStyle(
            'ReportTitle',
            parent=styles['Title'],
            fontName='Helvetica-Bold',
            fontSize=22,
            leading=26,
            textColor=colors.HexColor('#1e1b4b'), # Indigo 950
            alignment=0, # Left-aligned
            spaceAfter=20
        )
        
        h1_style = ParagraphStyle(
            'SectionH1',
            parent=styles['Heading1'],
            fontName='Helvetica-Bold',
            fontSize=15,
            leading=18,
            textColor=colors.HexColor('#4f46e5'), # Indigo 600
            spaceBefore=15,
            spaceAfter=10,
            keepWithNext=True
        )
        
        meta_style = ParagraphStyle(
            'MetaText',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=9,
            leading=12,
            textColor=colors.HexColor('#64748b'), # Slate 500
            spaceAfter=15
        )
        
        body_style = ParagraphStyle(
            'Body',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=10,
            leading=14,
            textColor=colors.HexColor('#334155'), # Slate 700
            spaceAfter=6
        )

        bold_style = ParagraphStyle(
            'BodyBold',
            parent=body_style,
            fontName='Helvetica-Bold'
        )

        story = []
        
        # Cover/Header Metadata
        story.append(Paragraph("QA Requirement Analysis & Test Case Report", title_style))
        source_label = str(data.get("source_type", "text")).upper()
        story.append(Paragraph(
            f"Source Analysis Type: <b>{source_label}</b> | "
            f"AI Confidence Score: <b>{int(data.get('confidence_score', 0) * 100)}%</b> | "
            f"Report Generated: <b>{datetime_str()}</b>",
            meta_style
        ))
        story.append(Spacer(1, 10))

        # --- 1. SCENARIOS ---
        story.append(Paragraph("1. Test Scenarios", h1_style))
        for item in data.get("scenarios", []):
            p_text = f"<b>{item.get('id', 'TS')}: {item.get('title', '')}</b><br/>{item.get('description', '')}"
            story.append(Paragraph(p_text, body_style))
            story.append(Spacer(1, 5))
        story.append(Spacer(1, 10))

        # --- 2. TEST CASES ---
        story.append(Paragraph("2. Test Cases", h1_style))
        for item in data.get("test_cases", []):
            story.append(Paragraph(f"<b>{item.get('id', 'TC')}: {item.get('title', '')}</b> (Ref: {item.get('scenario_id', '')})", bold_style))
            story.append(Paragraph(f"<i>Preconditions:</i> {item.get('preconditions', '')}", body_style))
            
            steps = item.get("steps", [])
            steps_html = "<br/>".join([f"{idx+1}. {step}" for idx, step in enumerate(steps)])
            story.append(Paragraph(f"<i>Execution Steps:</i><br/>{steps_html}", body_style))
            story.append(Paragraph(f"<i>Expected Result:</i> <b>{item.get('expected_result', '')}</b>", body_style))
            story.append(Spacer(1, 8))
        story.append(Spacer(1, 10))

        # --- 3. POSITIVE CASES ---
        story.append(Paragraph("3. Positive Cases", h1_style))
        for item in data.get("positive_cases", []):
            story.append(Paragraph(f"<b>{item.get('id', 'PC')}: {item.get('title', '')}</b>", bold_style))
            steps = item.get("steps", [])
            steps_html = "<br/>".join([f"- {step}" for step in steps])
            story.append(Paragraph(f"<i>Steps:</i><br/>{steps_html}", body_style))
            story.append(Paragraph(f"<i>Expected:</i> {item.get('expected_result', '')}", body_style))
            story.append(Spacer(1, 5))
        story.append(Spacer(1, 10))

        # --- 4. NEGATIVE CASES ---
        story.append(Paragraph("4. Negative Cases", h1_style))
        for item in data.get("negative_cases", []):
            story.append(Paragraph(f"<b>{item.get('id', 'NC')}: {item.get('title', '')}</b>", bold_style))
            steps = item.get("steps", [])
            steps_html = "<br/>".join([f"- {step}" for step in steps])
            story.append(Paragraph(f"<i>Steps:</i><br/>{steps_html}", body_style))
            story.append(Paragraph(f"<i>Expected:</i> {item.get('expected_result', '')}", body_style))
            story.append(Spacer(1, 5))
        story.append(Spacer(1, 10))

        # --- 5. EDGE CASES ---
        story.append(Paragraph("5. Edge Cases", h1_style))
        for item in data.get("edge_cases", []):
            story.append(Paragraph(f"<b>{item.get('id', 'EC')}: {item.get('title', '')}</b>", bold_style))
            steps = item.get("steps", [])
            steps_html = "<br/>".join([f"- {step}" for step in steps])
            story.append(Paragraph(f"<i>Steps:</i><br/>{steps_html}", body_style))
            story.append(Paragraph(f"<i>Expected:</i> {item.get('expected_result', '')}", body_style))
            story.append(Spacer(1, 5))
        story.append(Spacer(1, 10))

        # --- 6. RISKS ---
        story.append(Paragraph("6. Risks & Mitigations", h1_style))
        for item in data.get("risks", []):
            story.append(Paragraph(
                f"<b>{item.get('id', 'RK')}: {item.get('description', '')}</b><br/>"
                f"Impact: <b>{item.get('impact', '')}</b><br/>"
                f"Mitigation: {item.get('mitigation', '')}", 
                body_style
            ))
            story.append(Spacer(1, 6))
        story.append(Spacer(1, 10))

        # --- 7. MISSING REQUIREMENTS ---
        story.append(Paragraph("7. Missing Specifications & Ambiguities", h1_style))
        for item in data.get("missing_requirements", []):
            story.append(Paragraph(
                f"<b>{item.get('id', 'MR')}: {item.get('description', '')}</b> (Impact: {item.get('impact', '')})", 
                body_style
            ))
            story.append(Spacer(1, 5))

        # Build PDF
        doc.build(story)
        pdf_data = buffer.getvalue()
        buffer.close()
        return pdf_data

    @staticmethod
    def generate_excel_report(data: dict) -> bytes:
        """
        Generates a premium, multi-sheet Excel report from the analysis JSON.
        """
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        default_sheet = wb.active
        if default_sheet is not None:
            wb.remove(default_sheet)
            
        # Common premium styles
        header_fill = PatternFill(start_color="1e1b4b", end_color="1e1b4b", fill_type="solid") # Dark Indigo
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        title_font = Font(name="Calibri", size=14, bold=True, color="1e1b4b")
        data_font = Font(name="Calibri", size=11)
        bold_data_font = Font(name="Calibri", size=11, bold=True)
        zebra_fill = PatternFill(start_color="f8fafc", end_color="f8fafc", fill_type="solid") # Light gray-blue
        
        thin_border = Border(
            left=Side(style='thin', color='dddddd'),
            right=Side(style='thin', color='dddddd'),
            top=Side(style='thin', color='dddddd'),
            bottom=Side(style='thin', color='dddddd')
        )
        
        # Sheets map
        sheets_config = [
            ("Scenarios", ["ID", "Title", "Description"], ["id", "title", "description"]),
            ("Test Cases", ["ID", "Scenario ID", "Title", "Preconditions", "Steps", "Expected Result"], ["id", "scenario_id", "title", "preconditions", "steps", "expected_result"]),
            ("Positive Cases", ["ID", "Title", "Steps", "Expected Result"], ["id", "title", "steps", "expected_result"]),
            ("Negative Cases", ["ID", "Title", "Steps", "Expected Result"], ["id", "title", "steps", "expected_result"]),
            ("Edge Cases", ["ID", "Title", "Steps", "Expected Result"], ["id", "title", "steps", "expected_result"]),
            ("Risks", ["ID", "Risk Description", "Impact Level", "Recommended Mitigation"], ["id", "description", "impact", "mitigation"]),
            ("Missing Requirements", ["ID", "Description", "Impact Level"], ["id", "description", "impact"])
        ]

        for title, headers, keys in sheets_config:
            ws = wb.create_sheet(title=title)
            
            # Show grid lines
            ws.views.sheetView[0].showGridLines = True
            
            # Title block
            ws.cell(row=1, column=1, value=f"QA {title} Analysis Report").font = title_font
            ws.row_dimensions[1].height = 25
            
            # Headers
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                cell.border = thin_border
            ws.row_dimensions[3].height = 28
            
            # Populate Data
            items_list = []
            if title == "Scenarios":
                items_list = data.get("scenarios", [])
            elif title == "Test Cases":
                items_list = data.get("test_cases", [])
            elif title == "Positive Cases":
                items_list = data.get("positive_cases", [])
            elif title == "Negative Cases":
                items_list = data.get("negative_cases", [])
            elif title == "Edge Cases":
                items_list = data.get("edge_cases", [])
            elif title == "Risks":
                items_list = data.get("risks", [])
            elif title == "Missing Requirements":
                items_list = data.get("missing_requirements", [])

            for row_idx, item in enumerate(items_list, 4):
                ws.row_dimensions[row_idx].height = 20
                for col_idx, key in enumerate(keys, 1):
                    val = item.get(key, "")
                    if isinstance(val, list):
                        # Join lists (e.g., test steps) with newlines for excel formatting
                        val = "\n".join([f"{i+1}. {x}" if title == "Test Cases" else f"- {x}" for i, x in enumerate(val)])
                    
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.font = bold_data_font if col_idx == 1 else data_font
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                    
                    # Zebra striping
                    if row_idx % 2 == 0:
                        cell.fill = zebra_fill
            
            # Auto-fit column widths
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.row < 3: # Ignore title block row for size calculations
                        continue
                    if cell.value:
                        lines = str(cell.value).split('\n')
                        max_len = max(max_len, max(len(l) for l in lines))
                ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 45) # Keep width between 12 and 45

        buffer = io.BytesIO()
        wb.save(buffer)
        excel_data = buffer.getvalue()
        buffer.close()
        return excel_data

def datetime_str() -> str:
    import datetime
    return datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
